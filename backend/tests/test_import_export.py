"""Bulk import and Excel export."""

import io
import json
from datetime import datetime

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.tokens import create_access_token
from app.models.data_record import DataRecord, DataSource
from app.models.user import UserRole
from app.services import excel, importer
from tests.conftest import auth_headers

CSV_HEADER = "title,value,category,timestamp"
CSV_ROWS = [
    "CPU reading,42.5,cpu,2026-08-30T10:00:00",
    "Memory reading,73.2,memory,2026-08-30T10:00:01",
    "Disk reading,11.0,disk,2026-08-30T10:00:02",
]
VALID_CSV = ("\n".join([CSV_HEADER, *CSV_ROWS]) + "\n").encode()

JSON_ROWS = [
    {"title": "CPU reading", "value": 42.5, "category": "cpu", "timestamp": "2026-08-30T10:00:00"},
    {"title": "Memory reading", "value": 73.2, "category": "memory", "timestamp": "2026-08-30T10:00:01"},
]
VALID_JSON = json.dumps(JSON_ROWS).encode()


def token_for(user):
    return auth_headers(create_access_token(user.id))


async def upload(client, user, content: bytes, filename="records.csv"):
    return await client.post(
        "/records/import",
        headers=token_for(user),
        files={"file": (filename, content, "text/csv")},
    )


async def stored(db_session) -> list[DataRecord]:
    return list((await db_session.execute(select(DataRecord))).scalars().all())


# --------------------------------------------------------------------------
# CSV import
# --------------------------------------------------------------------------


async def test_valid_csv_imports_every_row(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    response = await upload(client, user, VALID_CSV)

    assert response.status_code == 201
    assert response.json() == {
        "imported": 3,
        "source": "IMPORT",
        "filename": "records.csv",
    }
    assert len(await stored(db_session)) == 3


async def test_imported_rows_are_marked_and_owned_by_the_importer(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    await upload(client, user, VALID_CSV)

    records = await stored(db_session)
    assert all(record.source is DataSource.IMPORT for record in records)
    assert all(record.owner_id == user.id for record in records)
    assert all(record.is_anomaly is False for record in records)


async def test_values_and_timestamps_survive_the_round_trip(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    await upload(client, user, VALID_CSV)

    records = sorted(await stored(db_session), key=lambda r: r.timestamp)
    assert [r.value for r in records] == [42.5, 73.2, 11.0]
    assert records[0].timestamp == datetime(2026, 8, 30, 10, 0, 0)
    assert records[0].title == "CPU reading"


async def test_blank_lines_are_skipped(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    content = (CSV_HEADER + "\n" + CSV_ROWS[0] + "\n\n\n" + CSV_ROWS[1] + "\n").encode()

    response = await upload(client, user, content)

    assert response.json()["imported"] == 2
    assert len(await stored(db_session)) == 2


async def test_utf8_content_is_accepted(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    content = f"{CSV_HEADER}\n溫度讀數,25.5,溫度,2026-08-30T10:00:00\n".encode()

    response = await upload(client, user, content)

    assert response.status_code == 201
    assert (await stored(db_session))[0].title == "溫度讀數"


async def test_missing_column_is_rejected(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    content = b"title,value,category\nCPU,1.0,cpu\n"

    response = await upload(client, user, content)

    assert response.status_code == 400
    assert "timestamp" in response.json()["detail"]
    assert await stored(db_session) == []


async def test_unsupported_column_is_rejected(client, make_user):
    """A file may not smuggle server-controlled fields in as columns."""
    user = await make_user(UserRole.USER)
    content = (CSV_HEADER + ",owner_id\n" + CSV_ROWS[0] + ",999\n").encode()

    response = await upload(client, user, content)

    assert response.status_code == 400
    assert "owner_id" in response.json()["detail"]


@pytest.mark.parametrize(
    ("row", "expected"),
    [
        ("CPU,not-a-number,cpu,2026-08-30T10:00:00", "value"),
        ("CPU,1.0,cpu,not-a-date", "timestamp"),
        (",1.0,cpu,2026-08-30T10:00:00", "title"),
        ("CPU,1.0,,2026-08-30T10:00:00", "category"),
    ],
)
async def test_invalid_row_is_rejected_with_its_position(client, make_user, row, expected):
    user = await make_user(UserRole.USER)
    content = (CSV_HEADER + "\n" + row + "\n").encode()

    response = await upload(client, user, content)

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail.startswith("Row 2:")
    assert expected in detail


async def test_import_is_atomic(client, db_session, make_user):
    """A late invalid row must leave the earlier ones unwritten."""
    user = await make_user(UserRole.USER)
    rows = [f"Reading {index},{index}.0,cpu,2026-08-30T10:00:00" for index in range(56)]
    rows.append("Broken,not-a-number,cpu,2026-08-30T10:00:00")
    rows += [f"Reading {index},{index}.0,cpu,2026-08-30T10:00:00" for index in range(57, 100)]
    content = ("\n".join([CSV_HEADER, *rows]) + "\n").encode()

    response = await upload(client, user, content)

    assert response.status_code == 400
    assert "Row 58" in response.json()["detail"]
    assert await stored(db_session) == []  # rows 1-56 were not inserted


# --------------------------------------------------------------------------
# JSON import
# --------------------------------------------------------------------------


async def test_valid_json_array_imports(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    response = await upload(client, user, VALID_JSON, "records.json")

    assert response.status_code == 201
    assert response.json()["imported"] == 2
    records = await stored(db_session)
    assert {r.category for r in records} == {"cpu", "memory"}
    assert all(r.source is DataSource.IMPORT for r in records)


async def test_malformed_json_is_rejected(client, db_session, make_user):
    user = await make_user(UserRole.USER)

    response = await upload(client, user, b"{not json", "records.json")

    assert response.status_code == 400
    assert "Invalid JSON" in response.json()["detail"]
    assert await stored(db_session) == []


@pytest.mark.parametrize("payload", [b'{"title": "x"}', b'"a string"', b"42", b"null"])
async def test_json_must_be_a_top_level_array(client, make_user, payload):
    user = await make_user(UserRole.USER)

    response = await upload(client, user, payload, "records.json")

    assert response.status_code == 400
    assert "top-level array" in response.json()["detail"]


async def test_json_entry_must_be_an_object(client, make_user):
    user = await make_user(UserRole.USER)

    response = await upload(client, user, b'[1, 2]', "records.json")

    assert response.status_code == 400
    assert "expected an object" in response.json()["detail"]


async def test_json_missing_field_is_rejected(client, make_user):
    user = await make_user(UserRole.USER)
    payload = json.dumps([{"title": "CPU", "value": 1.0, "category": "cpu"}]).encode()

    response = await upload(client, user, payload, "records.json")

    assert response.status_code == 400
    assert "timestamp" in response.json()["detail"]


@pytest.mark.parametrize(
    "forbidden",
    [
        {"owner_id": 999},
        {"is_anomaly": True},
        {"source": "REALTIME"},
        {"source": "MANUAL"},
        {"id": 1},
        {"created_at": "2026-08-30T10:00:00"},
        {"updated_at": "2026-08-30T10:00:00"},
    ],
)
async def test_system_fields_in_json_are_rejected(client, db_session, make_user, forbidden):
    """Forbidden fields are refused outright, never silently ignored."""
    user = await make_user(UserRole.USER)
    payload = json.dumps([{**JSON_ROWS[0], **forbidden}]).encode()

    response = await upload(client, user, payload, "records.json")

    assert response.status_code == 400
    assert list(forbidden)[0] in response.json()["detail"]
    assert await stored(db_session) == []


async def test_json_import_is_atomic(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    payload = json.dumps([*JSON_ROWS, {**JSON_ROWS[0], "value": "abc"}]).encode()

    response = await upload(client, user, payload, "records.json")

    assert response.status_code == 400
    assert await stored(db_session) == []


# --------------------------------------------------------------------------
# Formats, limits and permissions
# --------------------------------------------------------------------------


@pytest.mark.parametrize("filename", ["records.txt", "records.xlsx", "records", "records.csv.exe"])
async def test_unsupported_extension_is_rejected(client, make_user, filename):
    user = await make_user(UserRole.USER)

    response = await upload(client, user, VALID_CSV, filename)

    assert response.status_code == 400
    assert ".csv and .json" in response.json()["detail"]


@pytest.mark.parametrize("content", [b"", b"title,value,category,timestamp\n", b"[]"])
async def test_empty_import_is_rejected(client, make_user, content):
    user = await make_user(UserRole.USER)
    name = "records.json" if content == b"[]" else "records.csv"

    response = await upload(client, user, content, name)

    assert response.status_code == 400
    assert "no records" in response.json()["detail"] or "empty" in response.json()["detail"]


async def test_too_many_records_is_rejected(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    rows = [f"Reading,{index}.0,cpu,2026-08-30T10:00:00" for index in range(importer.MAX_IMPORT_RECORDS + 1)]
    content = ("\n".join([CSV_HEADER, *rows]) + "\n").encode()

    response = await upload(client, user, content)

    assert response.status_code == 400
    assert str(importer.MAX_IMPORT_RECORDS) in response.json()["detail"]
    assert await stored(db_session) == []


def test_oversized_file_is_rejected_before_parsing():
    oversized = b"x" * (importer.MAX_IMPORT_BYTES + 1)

    with pytest.raises(importer.ImportError_, match="import limit"):
        importer.parse_upload("records.csv", oversized)


async def test_viewer_may_not_import(client, db_session, make_user):
    viewer = await make_user(UserRole.VIEWER)

    response = await upload(client, viewer, VALID_CSV)

    assert response.status_code == 403
    assert await stored(db_session) == []


async def test_admin_may_import(client, make_user):
    admin = await make_user(UserRole.ADMIN)

    assert (await upload(client, admin, VALID_CSV)).status_code == 201


async def test_unauthenticated_import_is_rejected(client):
    response = await client.post(
        "/records/import", files={"file": ("records.csv", VALID_CSV, "text/csv")}
    )

    assert response.status_code == 401


# --------------------------------------------------------------------------
# Excel export
# --------------------------------------------------------------------------


async def seed(db_session, count=3, **overrides):
    for index in range(count):
        db_session.add(
            DataRecord(
                title=overrides.get("title", f"Reading {index}"),
                value=float(index) + 0.5,
                category=overrides.get("category", "cpu"),
                timestamp=datetime(2026, 8, 30, 10, index, 0),
                source=overrides.get("source", DataSource.MANUAL),
                is_anomaly=index == 0,
                owner_id=overrides.get("owner_id"),
            )
        )
    await db_session.commit()


def workbook_from(response):
    return load_workbook(io.BytesIO(response.content))


async def export(client, user, **params):
    return await client.get("/records/export.xlsx", headers=token_for(user), params=params)


async def test_export_returns_an_xlsx_attachment(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    await seed(db_session)

    response = await export(client, user)

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == 'attachment; filename="data_records.xlsx"'
    assert response.content[:2] == b"PK"  # a real zip-based workbook


async def test_workbook_opens_with_expected_headers_and_rows(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    await seed(db_session)

    sheet = workbook_from(await export(client, user)).active

    assert sheet.title == "Data records"
    assert [cell.value for cell in sheet[1]] == [
        "ID", "Title", "Value", "Category", "Timestamp",
        "Source", "Anomaly", "Owner ID", "Created at", "Updated at",
    ]
    assert sheet.max_row == 4  # header + 3 records
    assert [row[1].value for row in sheet.iter_rows(min_row=2)] == [
        "Reading 0", "Reading 1", "Reading 2"
    ]


async def test_values_stay_numeric_and_timestamps_stay_dates(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    await seed(db_session)

    sheet = workbook_from(await export(client, user)).active
    first = list(sheet.iter_rows(min_row=2, max_row=2))[0]

    assert isinstance(first[2].value, float) and first[2].value == 0.5
    assert isinstance(first[4].value, datetime)
    assert first[5].value == "MANUAL"
    assert first[6].value is True  # anomaly flag as a boolean


async def test_export_honours_filters(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    await seed(db_session, 2, category="cpu")
    await seed(db_session, 3, category="memory", source=DataSource.REALTIME)

    by_category = workbook_from(await export(client, user, category="memory")).active
    by_source = workbook_from(await export(client, user, source="REALTIME")).active
    ranged = workbook_from(
        await export(client, user, start="2026-08-30T10:01:00", end="2026-08-30T10:02:00")
    ).active

    assert by_category.max_row == 4
    assert by_source.max_row == 4
    assert ranged.max_row >= 2


async def test_empty_result_produces_a_header_only_workbook(client, make_user):
    user = await make_user(UserRole.USER)

    response = await export(client, user, category="nothing-here")
    sheet = workbook_from(response).active

    assert response.status_code == 200
    assert sheet.max_row == 1
    assert sheet["A1"].value == "ID"
    assert response.headers["x-export-rows"] == "0"


async def test_export_reports_its_row_limit(client, db_session, make_user):
    user = await make_user(UserRole.USER)
    await seed(db_session)

    response = await export(client, user)

    assert response.headers["x-export-rows"] == "3"
    assert response.headers["x-export-truncated"] == "false"
    assert int(response.headers["x-export-row-limit"]) >= 1000


@pytest.mark.parametrize("role", [UserRole.ADMIN, UserRole.USER, UserRole.VIEWER])
async def test_every_role_may_export(client, db_session, make_user, role):
    user = await make_user(role)
    await seed(db_session, 1)

    assert (await export(client, user)).status_code == 200


async def test_unauthenticated_export_is_rejected(client):
    assert (await client.get("/records/export.xlsx")).status_code == 401


async def test_invalid_export_filter_is_rejected(client, make_user):
    user = await make_user(UserRole.USER)

    assert (await export(client, user, source="NOT_A_SOURCE")).status_code == 422
    assert (await export(client, user, start="not-a-date")).status_code == 422


# --------------------------------------------------------------------------
# Formula injection
# --------------------------------------------------------------------------


@pytest.mark.parametrize("dangerous", ["=1+1", "+1", "-1", "@SUM(A1)", '=cmd|"/c calc"!A1'])
def test_formula_like_text_is_neutralised(dangerous):
    assert excel.neutralise(dangerous) == f"'{dangerous}"


@pytest.mark.parametrize("safe", ["CPU reading", "cpu", "reading =1", "", "1+1"])
def test_ordinary_text_is_left_alone(safe):
    assert excel.neutralise(safe) == safe


async def test_formula_text_from_an_import_is_neutralised_in_the_workbook(
    client, db_session, make_user
):
    user = await make_user(UserRole.USER)
    await seed(db_session, 1, title="=1+1", category="@evil")

    sheet = workbook_from(await export(client, user)).active
    row = list(sheet.iter_rows(min_row=2, max_row=2))[0]

    assert row[1].value == "'=1+1"
    assert row[3].value == "'@evil"
